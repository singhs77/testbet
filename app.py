from flask import Flask, request, jsonify
from flask import Flask, request, redirect, url_for, render_template
app = Flask(__name__)

from flask_cors import CORS
from PIL import Image
import pytesseract
import openpyxl
import os
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
import re




def parse_extracted_text(extracted_text):
    # Define patterns for each field
    team_name_pattern = r"^([A-Za-z\s]+) [+-]?\d+"  # Matches "Tennessee +245" but not "MONEYLINE"
    opponent_pattern = r"@ ([A-Za-z\s]+)"  # Matches "@ Georgia"
    moneyline_pattern = r"MONEYLINE"       # Identifies moneyline bet type
    wager_pattern = r"Wager \$([\d,]+\.\d{2})"  # Matches "Wager $250.00"
    odds_pattern = r"Odds ([+-]?\d+)"          # Matches "Odds -106" or "Odds +245"
    to_win_pattern = r"To win \$([\d,]+\.\d{2})"  # Matches "To win $236.90"
    total_payout_pattern = r"Total payout \$([\d,]+\.\d{2})"  # Matches "Total payout $486.90"

    # Default values
    team_name = None
    opponent = "Parlay"  # Default for parlay bets
    bet_type = "Parlay"  # Default to "Parlay"
    wagered_amount = None
    odds = None
    to_win = None
    total_payout = None

    # Extract team name
    team_name_match = re.search(team_name_pattern, extracted_text, re.MULTILINE)
    if team_name_match:
        team_name = team_name_match.group(1).strip()

    # Extract opponent (for moneyline bets)
    opponent_match = re.search(opponent_pattern, extracted_text)
    if opponent_match:
        opponent = opponent_match.group(1).strip()

    # Check if it's a moneyline bet
    if re.search(moneyline_pattern, extracted_text, re.MULTILINE):
        bet_type = "Moneyline"

    # Extract wagered amount
    wager_match = re.search(wager_pattern, extracted_text, re.MULTILINE)
    if wager_match:
        wagered_amount = f"${wager_match.group(1)}"

    # Extract odds
    odds_match = re.search(odds_pattern, extracted_text, re.MULTILINE)
    if odds_match:
        odds = odds_match.group(1)

    # Extract "To Win" amount
    to_win_match = re.search(to_win_pattern, extracted_text, re.MULTILINE)
    if to_win_match:
        to_win = f"${to_win_match.group(1)}"

    # Extract total payout
    total_payout_match = re.search(total_payout_pattern, extracted_text, re.MULTILINE)
    if total_payout_match:
        total_payout = f"${total_payout_match.group(1)}"

    # Return parsed data
    return {
        "team_name": team_name,
        "opponent": opponent,
        "bet_type": bet_type,
        "wagered_amount": wagered_amount,
        "odds": odds,
        "to_win": to_win,
        "total_payout": total_payout,
    }



app = Flask(__name__)
CORS(app)  # Allow cross-origin requests from your HTML

EXCEL_FILE = 'betslip4_data.xlsx'

# Create the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Team", "Opponent", "BetType", "Wagered Amount", "Odds", "To Win", "Total Payout"])  # Example headers
    workbook.save(EXCEL_FILE)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    # Retrieve form data
    name = request.form.get('name')
    date = request.form.get('date')
    description = request.form.get('description')

    file = request.files['file']

    try:
        # Open the uploaded image
        image = Image.open(file)

        # Extract text using Tesseract OCR
        extracted_text = pytesseract.image_to_string(image)
        print("Extracted Text from Image:")
        print(extracted_text)

        # Parse extracted text
        parsed_data = parse_extracted_text(extracted_text)
        print("Parsed Data:", parsed_data)

        # Prepare row for Excel
        data_row = [
            name,
            date,
            description,
            parsed_data["team_name"],
            parsed_data["opponent"],
            parsed_data["bet_type"],
            parsed_data["wagered_amount"],
            parsed_data["odds"],
            parsed_data["to_win"],
            parsed_data["total_payout"],
        ]
        print(f"Row for Excel: {data_row}")

        # Save data to Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # Add headers if needed
        if sheet.max_row == 1:
            headers = [
                "Name", "Date", "Description", "Team Name", "Opponent",
                "Bet Type", "Wagered Amount", "Odds", "To Win", "Total Payout"
            ]
            sheet.append(headers)

        # Append the row
        sheet.append(data_row)
        workbook.save(EXCEL_FILE)

        return jsonify({"message": "Data successfully added to Excel"})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)
